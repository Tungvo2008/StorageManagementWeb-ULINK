from decimal import Decimal
from io import BytesIO
from pathlib import Path
import re
from typing import Any
import unicodedata
from uuid import uuid4

from fastapi import APIRouter, Depends, HTTPException, Response, UploadFile, File, status
from fastapi.responses import FileResponse
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.api.deps import get_db
from app.core.config import settings
from app.db.models import AmazonProductMapping, Category, Product
from app.schemas.product import ProductCreate, ProductImportResult, ProductRead, ProductUpdate


router = APIRouter(prefix="/products")

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PRODUCT_IMAGE_DIR = Path(__file__).resolve().parents[3] / "assets" / "product-images"
PRODUCT_IMAGE_URL_PREFIX = f"{settings.API_V1_STR}/public/product-images/"
LEGACY_PRODUCT_IMAGE_URL_PREFIX = f"{settings.API_V1_STR}/products/image-files/"
MAX_IMAGE_BYTES = 10 * 1024 * 1024
ALLOWED_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
CONTENT_TYPE_TO_SUFFIX = {
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/webp": ".webp",
    "image/gif": ".gif",
}


def _has_python_multipart() -> bool:
    try:
        import multipart  # type: ignore  # noqa: F401
    except Exception:
        return False
    return True


def _openpyxl():
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore
        from openpyxl.styles import Font  # type: ignore
    except Exception as e:
        raise RuntimeError("Missing dependency: openpyxl. Please `pip install openpyxl`.") from e
    return Workbook, load_workbook, Font


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _parse_bool(v: Any, default: bool = True) -> bool:
    s = _cell_str(v).lower()
    if not s:
        return default
    return s in {"1", "true", "yes", "y", "active"}


def _parse_decimal(v: Any, *, default: Decimal = Decimal("0")) -> Decimal:
    raw = _cell_str(v)
    if not raw:
        return default
    normalized = raw.replace(",", "")
    try:
        return Decimal(normalized)
    except Exception as e:
        raise ValueError(f"Invalid decimal value: {raw}") from e


def _ensure_product_image_dir() -> None:
    PRODUCT_IMAGE_DIR.mkdir(parents=True, exist_ok=True)


def _detect_image_suffix(upload: UploadFile) -> str:
    suffix = Path(upload.filename or "").suffix.lower()
    if suffix in ALLOWED_IMAGE_SUFFIXES:
        return ".jpg" if suffix == ".jpeg" else suffix
    mapped = CONTENT_TYPE_TO_SUFFIX.get((upload.content_type or "").lower())
    if mapped:
        return mapped
    raise HTTPException(status_code=400, detail="Unsupported image type. Use png, jpg, webp, or gif.")


def _delete_managed_product_image(image_url: str | None) -> None:
    if not image_url:
        return
    matching_prefix = next(
        (
            prefix
            for prefix in (PRODUCT_IMAGE_URL_PREFIX, LEGACY_PRODUCT_IMAGE_URL_PREFIX)
            if image_url.startswith(prefix)
        ),
        None,
    )
    if matching_prefix is None:
        return
    filename = Path(image_url.removeprefix(matching_prefix)).name
    if not filename:
        return
    file_path = PRODUCT_IMAGE_DIR / filename
    if file_path.exists():
        file_path.unlink()


def _resolve_managed_product_image_path(filename: str) -> Path:
    safe_name = Path(filename).name
    if not safe_name:
        raise HTTPException(status_code=404, detail="Image not found")
    file_path = PRODUCT_IMAGE_DIR / safe_name
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Image not found")
    return file_path


def _normalize_sku_value(value: str | None) -> str:
    return (value or "").strip()


def _slugify_sku_from_name(name: str) -> str:
    normalized = unicodedata.normalize("NFKD", name or "")
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = ascii_text.upper()
    ascii_text = re.sub(r"[^A-Z0-9]+", "_", ascii_text)
    ascii_text = re.sub(r"_+", "_", ascii_text).strip("_")
    return ascii_text[:64] or "SAN_PHAM"


def _ensure_unique_sku(base_sku: str, existing_skus_upper: set[str]) -> str:
    candidate = base_sku[:64] or "SAN_PHAM"
    if candidate.upper() not in existing_skus_upper:
        return candidate
    suffix = 2
    while True:
        suffix_str = f"_{suffix}"
        trimmed = candidate[: max(1, 64 - len(suffix_str))]
        next_candidate = f"{trimmed}{suffix_str}"
        if next_candidate.upper() not in existing_skus_upper:
            return next_candidate
        suffix += 1


def _resolve_sku(
    *,
    raw_sku: str | None,
    name: str,
    existing_skus_upper: set[str],
    current_sku_upper: str | None = None,
) -> str:
    explicit_sku = _normalize_sku_value(raw_sku)
    if explicit_sku:
        if explicit_sku.upper() != (current_sku_upper or "") and explicit_sku.upper() in existing_skus_upper:
            raise HTTPException(status_code=409, detail="SKU already exists")
        return explicit_sku

    generated = _slugify_sku_from_name(name)
    if current_sku_upper and current_sku_upper == generated.upper():
        return generated
    taken = set(existing_skus_upper)
    if current_sku_upper:
        taken.discard(current_sku_upper)
    return _ensure_unique_sku(generated, taken)


def _prepare_amazon_fields(
    db: Session,
    *,
    is_sold_on_amazon: bool,
    amazon_sku: str | None,
    current_product_id: int | None = None,
) -> tuple[bool, str | None]:
    if not is_sold_on_amazon:
        return False, None
    normalized_sku = (amazon_sku or "").strip()
    if not normalized_sku:
        raise HTTPException(status_code=400, detail="Amazon SKU is required when the product is sold on Amazon")
    if len(normalized_sku) > 80:
        raise HTTPException(status_code=400, detail="Amazon SKU must be 80 characters or fewer")
    if not normalized_sku.isascii():
        raise HTTPException(status_code=400, detail="Amazon SKU must use English/ASCII characters")

    existing_product = db.scalar(select(Product).where(Product.amazon_sku == normalized_sku))
    if existing_product is not None and existing_product.id != current_product_id:
        raise HTTPException(status_code=409, detail="Amazon SKU is already assigned to another product")
    existing_mapping = db.scalar(
        select(AmazonProductMapping).where(AmazonProductMapping.amazon_sku == normalized_sku)
    )
    if (
        existing_mapping is not None
        and existing_mapping.product_id is not None
        and existing_mapping.product_id != current_product_id
    ):
        raise HTTPException(status_code=409, detail="Amazon SKU is already mapped to another product")
    return True, normalized_sku


def _sync_product_amazon_mapping(db: Session, product: Product) -> None:
    if not product.is_sold_on_amazon or not product.amazon_sku:
        return
    target = db.scalar(
        select(AmazonProductMapping).where(AmazonProductMapping.amazon_sku == product.amazon_sku)
    )
    if target is not None:
        if target.product_id not in (None, product.id):
            raise HTTPException(status_code=409, detail="Amazon SKU is already mapped to another product")
        target.product_id = product.id
        if not target.title:
            target.title = product.name
        return

    current = db.scalar(
        select(AmazonProductMapping)
        .where(AmazonProductMapping.product_id == product.id)
        .order_by(AmazonProductMapping.id.asc())
    )
    if current is not None:
        current.amazon_sku = product.amazon_sku
        if not current.title:
            current.title = product.name
        return

    db.add(
        AmazonProductMapping(
            product_id=product.id,
            amazon_sku=product.amazon_sku,
            title=product.name,
        )
    )


class _ProductExcelImportError(ValueError):
    def __init__(self, errors: list[str]):
        super().__init__("\n".join(errors))
        self.errors = errors


def _build_products_template_xlsx() -> bytes:
    Workbook, _, Font = _openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Products Import"
    headers = [
        "sku",
        "name",
        "category",
        "base_uom",
        "uom",
        "unit_price",
        "cost_price",
        "is_active",
        "brand",
        "catalog_short_name",
        "unit_size",
        "catalog_case_pack",
        "country_of_origin",
        "upc",
        "catalog_badges",
        "catalog_enabled",
        "catalog_sort_order",
        "is_sold_on_amazon",
        "amazon_sku",
    ]
    ws.append(headers)
    ws.append(
        [
            "",
            "Sample Product",
            "",
            "Pc",
            "Pc",
            "10.00",
            "6.50",
            True,
            "Sample Brand",
            "",
            "5 oz (142 g)",
            12,
            "Vietnam",
            "",
            "new,best seller",
            True,
            0,
            False,
            "",
        ]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 32
    ws.column_dimensions["K"].width = 18
    ws.column_dimensions["L"].width = 18
    ws.column_dimensions["M"].width = 20
    ws.column_dimensions["N"].width = 18
    ws.column_dimensions["O"].width = 24
    ws.column_dimensions["P"].width = 18
    ws.column_dimensions["Q"].width = 18
    ws.column_dimensions["R"].width = 20
    ws.column_dimensions["S"].width = 28

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_products_export_xlsx(products: list[Product]) -> bytes:
    Workbook, _, Font = _openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(
        [
            "id",
            "sku",
            "name",
            "category",
            "base_uom",
            "uom",
            "unit_price",
            "cost_price",
            "currency",
            "is_active",
            "brand",
            "catalog_short_name",
            "unit_size",
            "catalog_case_pack",
            "country_of_origin",
            "upc",
            "catalog_badges",
            "catalog_enabled",
            "catalog_sort_order",
            "is_sold_on_amazon",
            "amazon_sku",
        ]
    )
    for p in products:
        ws.append(
            [
                p.id,
                p.sku,
                p.name,
                p.category.name if p.category else "",
                p.base_uom,
                p.uom,
                str(p.unit_price or Decimal("0")),
                str(p.cost_price or Decimal("0")),
                p.currency or settings.DEFAULT_CURRENCY,
                bool(p.is_active),
                p.brand or "",
                p.catalog_short_name or "",
                p.unit_size or "",
                p.catalog_case_pack,
                p.country_of_origin or "",
                p.upc or "",
                p.catalog_badges or "",
                bool(p.catalog_enabled),
                int(p.catalog_sort_order or 0),
                bool(p.is_sold_on_amazon),
                p.amazon_sku or "",
            ]
        )
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 22
    ws.column_dimensions["L"].width = 32
    ws.column_dimensions["M"].width = 18
    ws.column_dimensions["N"].width = 18
    ws.column_dimensions["O"].width = 20
    ws.column_dimensions["P"].width = 18
    ws.column_dimensions["Q"].width = 24
    ws.column_dimensions["R"].width = 18
    ws.column_dimensions["S"].width = 18
    ws.column_dimensions["T"].width = 20
    ws.column_dimensions["U"].width = 28

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_products_import_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    _, load_workbook, _ = _openpyxl()
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise _ProductExcelImportError(["Empty workbook"])

    header = [(_cell_str(x) or "").strip().lower() for x in (rows[0] or [])]
    if "sku" not in header:
        raise _ProductExcelImportError(["Missing column: sku"])
    if "name" not in header:
        raise _ProductExcelImportError(["Missing column: name"])

    idx = {name: header.index(name) for name in header}
    errors: list[str] = []
    items: list[dict[str, Any]] = []
    for row_i, row in enumerate(rows[1:], start=2):
        row_vals = list(row or [])
        sku = _cell_str(row_vals[idx["sku"]] if idx["sku"] < len(row_vals) else "")
        name = _cell_str(row_vals[idx["name"]] if idx["name"] < len(row_vals) else "")
        if not sku and not name:
            continue
        if not name:
            errors.append(f"Row {row_i}: missing name")
            continue

        base_uom = _cell_str(row_vals[idx["base_uom"]] if "base_uom" in idx and idx["base_uom"] < len(row_vals) else "")
        uom = _cell_str(row_vals[idx["uom"]] if "uom" in idx and idx["uom"] < len(row_vals) else "")
        category_name = _cell_str(row_vals[idx["category"]] if "category" in idx and idx["category"] < len(row_vals) else "")
        unit_price_raw = row_vals[idx["unit_price"]] if "unit_price" in idx and idx["unit_price"] < len(row_vals) else None
        cost_price_raw = row_vals[idx["cost_price"]] if "cost_price" in idx and idx["cost_price"] < len(row_vals) else None
        is_active_raw = row_vals[idx["is_active"]] if "is_active" in idx and idx["is_active"] < len(row_vals) else None
        brand = _cell_str(row_vals[idx["brand"]] if "brand" in idx and idx["brand"] < len(row_vals) else "")
        catalog_short_name = _cell_str(
            row_vals[idx["catalog_short_name"]]
            if "catalog_short_name" in idx and idx["catalog_short_name"] < len(row_vals)
            else ""
        )
        unit_size = _cell_str(row_vals[idx["unit_size"]] if "unit_size" in idx and idx["unit_size"] < len(row_vals) else "")
        catalog_case_pack_raw = (
            row_vals[idx["catalog_case_pack"]]
            if "catalog_case_pack" in idx and idx["catalog_case_pack"] < len(row_vals)
            else None
        )
        country_of_origin = _cell_str(
            row_vals[idx["country_of_origin"]]
            if "country_of_origin" in idx and idx["country_of_origin"] < len(row_vals)
            else ""
        )
        upc = _cell_str(row_vals[idx["upc"]] if "upc" in idx and idx["upc"] < len(row_vals) else "")
        catalog_badges = _cell_str(
            row_vals[idx["catalog_badges"]]
            if "catalog_badges" in idx and idx["catalog_badges"] < len(row_vals)
            else ""
        )
        catalog_enabled_raw = (
            row_vals[idx["catalog_enabled"]]
            if "catalog_enabled" in idx and idx["catalog_enabled"] < len(row_vals)
            else None
        )
        catalog_sort_order_raw = (
            row_vals[idx["catalog_sort_order"]]
            if "catalog_sort_order" in idx and idx["catalog_sort_order"] < len(row_vals)
            else None
        )
        is_sold_on_amazon_raw = (
            row_vals[idx["is_sold_on_amazon"]]
            if "is_sold_on_amazon" in idx and idx["is_sold_on_amazon"] < len(row_vals)
            else None
        )
        amazon_sku = _cell_str(
            row_vals[idx["amazon_sku"]]
            if "amazon_sku" in idx and idx["amazon_sku"] < len(row_vals)
            else ""
        )

        base_uom = base_uom or "Pc"
        uom = uom or "Pc"
        is_active = _parse_bool(is_active_raw, default=True)
        catalog_enabled = _parse_bool(catalog_enabled_raw, default=True)
        is_sold_on_amazon = _parse_bool(
            is_sold_on_amazon_raw,
            default=bool(amazon_sku),
        )
        if is_sold_on_amazon and not amazon_sku:
            errors.append(f"Row {row_i}: amazon_sku is required when is_sold_on_amazon is true")
            continue
        if amazon_sku and (len(amazon_sku) > 80 or not amazon_sku.isascii()):
            errors.append(f"Row {row_i}: amazon_sku must be 80 or fewer ASCII characters")
            continue
        if not is_sold_on_amazon:
            amazon_sku = ""
        try:
            catalog_case_pack = (
                int(catalog_case_pack_raw)
                if catalog_case_pack_raw not in (None, "")
                else None
            )
            if catalog_case_pack is not None and catalog_case_pack < 1:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f"Row {row_i}: invalid catalog_case_pack")
            continue
        try:
            catalog_sort_order = int(catalog_sort_order_raw or 0)
            if catalog_sort_order < 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f"Row {row_i}: invalid catalog_sort_order")
            continue
        try:
            unit_price = _parse_decimal(unit_price_raw)
        except ValueError:
            errors.append(f"Row {row_i}: invalid unit_price")
            continue
        try:
            cost_price = _parse_decimal(cost_price_raw)
        except ValueError:
            errors.append(f"Row {row_i}: invalid cost_price")
            continue

        items.append(
            {
                "sku": sku,
                "name": name,
                "category": category_name,
                "base_uom": base_uom,
                "uom": uom,
                "unit_price": unit_price,
                "cost_price": cost_price,
                "is_active": is_active,
                "brand": brand,
                "catalog_short_name": catalog_short_name,
                "unit_size": unit_size,
                "catalog_case_pack": catalog_case_pack,
                "country_of_origin": country_of_origin,
                "upc": upc,
                "catalog_badges": catalog_badges,
                "catalog_enabled": catalog_enabled,
                "catalog_sort_order": catalog_sort_order,
                "is_sold_on_amazon": is_sold_on_amazon,
                "amazon_sku": amazon_sku,
                "_catalog_fields_present": {
                    field_name
                    for field_name in (
                        "brand",
                        "catalog_short_name",
                        "unit_size",
                        "catalog_case_pack",
                        "country_of_origin",
                        "upc",
                        "catalog_badges",
                        "catalog_enabled",
                        "catalog_sort_order",
                    )
                    if field_name in idx
                },
                "_amazon_fields_present": bool(
                    {"is_sold_on_amazon", "amazon_sku"}.intersection(idx)
                ),
            }
        )

    if errors:
        raise _ProductExcelImportError(errors)
    if not items:
        raise _ProductExcelImportError(["No product rows found"])
    return items


@router.get("", response_model=list[ProductRead])
def list_products(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)) -> list[Product]:
    stmt = select(Product).options(selectinload(Product.category)).offset(skip).limit(limit)
    return db.scalars(stmt).all()


@router.post("", response_model=ProductRead, status_code=status.HTTP_201_CREATED)
def create_product(product_in: ProductCreate, db: Session = Depends(get_db)) -> Product:
    existing_skus_upper = {sku.upper() for sku in db.scalars(select(Product.sku)).all() if sku}
    sku = _resolve_sku(raw_sku=product_in.sku, name=product_in.name, existing_skus_upper=existing_skus_upper)

    category_id = product_in.category_id
    if category_id is not None:
        cat = db.get(Category, category_id)
        if cat is None:
            raise HTTPException(status_code=400, detail="Invalid category_id")

    currency = product_in.currency or settings.DEFAULT_CURRENCY
    base_uom = (product_in.base_uom or "Pc").strip() or "Pc"
    uom = (product_in.uom or "Pc").strip() or "Pc"
    uom_multiplier = product_in.uom_multiplier
    if uom_multiplier is None:
        uom_multiplier = 12 if uom.lower() == "dozen" else 1
    if uom_multiplier <= 1:
        uom_multiplier = 1
        uom = base_uom
    is_sold_on_amazon, amazon_sku = _prepare_amazon_fields(
        db,
        is_sold_on_amazon=product_in.is_sold_on_amazon,
        amazon_sku=product_in.amazon_sku,
    )
    product = Product(
        category_id=category_id,
        sku=sku,
        name=product_in.name,
        description=product_in.description,
        image_url=product_in.image_url,
        brand=product_in.brand,
        catalog_short_name=product_in.catalog_short_name,
        unit_size=product_in.unit_size,
        catalog_case_pack=product_in.catalog_case_pack,
        country_of_origin=product_in.country_of_origin,
        upc=product_in.upc,
        catalog_badges=product_in.catalog_badges,
        catalog_enabled=product_in.catalog_enabled,
        catalog_sort_order=product_in.catalog_sort_order,
        is_sold_on_amazon=is_sold_on_amazon,
        amazon_sku=amazon_sku,
        base_uom=base_uom,
        uom=uom,
        uom_multiplier=uom_multiplier,
        cost_price=product_in.cost_price or Decimal("0"),
        unit_price=product_in.unit_price or Decimal("0"),
        currency=currency,
        quantity_on_hand=product_in.quantity_on_hand,
        is_active=product_in.is_active,
    )
    db.add(product)
    db.flush()
    _sync_product_amazon_mapping(db, product)
    db.commit()
    db.refresh(product)
    return product


@router.get("/template.xlsx")
def download_products_template() -> Response:
    try:
        xlsx = _build_products_template_xlsx()
    except RuntimeError as e:
        raise HTTPException(status_code=501, detail=str(e)) from e
    return Response(
        content=xlsx,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": 'attachment; filename="products-import-template.xlsx"'},
    )


@router.get("/export.xlsx")
def export_products_xlsx(db: Session = Depends(get_db)) -> Response:
    products = db.scalars(select(Product).options(selectinload(Product.category)).order_by(Product.id.asc())).all()
    try:
        xlsx = _build_products_export_xlsx(products)
    except RuntimeError as e:
        raise HTTPException(status_code=501, detail=str(e)) from e
    return Response(
        content=xlsx,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": 'attachment; filename="products-export.xlsx"'},
    )


if _has_python_multipart():

    @router.post("/import", response_model=ProductImportResult)
    async def import_products_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db)) -> ProductImportResult:
        content = await file.read()
        try:
            items = _parse_products_import_xlsx(content)
        except _ProductExcelImportError as e:
            raise HTTPException(status_code=400, detail="\n".join(e.errors)) from e
        except RuntimeError as e:
            raise HTTPException(status_code=501, detail=str(e)) from e

        existing_products = db.scalars(select(Product)).all()
        existing_by_sku = {p.sku.upper(): p for p in existing_products}
        categories = db.scalars(select(Category)).all()
        category_by_name = {cat.name.strip().lower(): cat for cat in categories if (cat.name or "").strip()}
        created = 0
        updated = 0
        for item in items:
            name = item["name"].strip()
            explicit_sku = _normalize_sku_value(item["sku"])
            sku = explicit_sku or _resolve_sku(
                raw_sku=None,
                name=name,
                existing_skus_upper=set(existing_by_sku.keys()),
            )
            key = sku.upper()
            base_uom = item["base_uom"].strip() or "Pc"
            uom = item["uom"].strip() or "Pc"
            uom_multiplier = 12 if uom.lower() == "dozen" else 1
            if uom_multiplier <= 1:
                uom_multiplier = 1
                uom = base_uom
            is_active = bool(item["is_active"])
            category_name = item["category"].strip()
            category = category_by_name.get(category_name.lower()) if category_name else None
            category_id = category.id if category is not None else None
            should_update_category = not category_name or category is not None
            unit_price = item["unit_price"]
            cost_price = item["cost_price"]
            all_catalog_fields = {
                "brand": item["brand"] or None,
                "catalog_short_name": item["catalog_short_name"] or None,
                "unit_size": item["unit_size"] or None,
                "catalog_case_pack": item["catalog_case_pack"],
                "country_of_origin": item["country_of_origin"] or None,
                "upc": item["upc"] or None,
                "catalog_badges": item["catalog_badges"] or None,
                "catalog_enabled": item["catalog_enabled"],
                "catalog_sort_order": item["catalog_sort_order"],
            }
            catalog_fields = {
                field_name: field_value
                for field_name, field_value in all_catalog_fields.items()
                if field_name in item["_catalog_fields_present"]
            }

            product = existing_by_sku.get(key)
            amazon_fields: dict[str, object] = {}
            if item["_amazon_fields_present"]:
                is_sold_on_amazon, amazon_sku = _prepare_amazon_fields(
                    db,
                    is_sold_on_amazon=bool(item["is_sold_on_amazon"]),
                    amazon_sku=item["amazon_sku"] or None,
                    current_product_id=product.id if product is not None else None,
                )
                amazon_fields = {
                    "is_sold_on_amazon": is_sold_on_amazon,
                    "amazon_sku": amazon_sku,
                }
            if product is None:
                product = Product(
                    sku=sku,
                    name=name,
                    category_id=category_id,
                    base_uom=base_uom,
                    uom=uom,
                    uom_multiplier=uom_multiplier,
                    is_active=is_active,
                    quantity_on_hand=0,
                    currency=settings.DEFAULT_CURRENCY,
                    unit_price=unit_price,
                    cost_price=cost_price,
                    **catalog_fields,
                    **amazon_fields,
                )
                db.add(product)
                db.flush()
                existing_by_sku[key] = product
                created += 1
            else:
                product.name = name
                if should_update_category:
                    product.category_id = category_id
                product.base_uom = base_uom
                product.uom = uom
                product.uom_multiplier = uom_multiplier
                product.unit_price = unit_price
                product.cost_price = cost_price
                product.is_active = is_active
                for field_name, field_value in catalog_fields.items():
                    setattr(product, field_name, field_value)
                for field_name, field_value in amazon_fields.items():
                    setattr(product, field_name, field_value)
                updated += 1
            if amazon_fields:
                _sync_product_amazon_mapping(db, product)

        db.commit()
        return ProductImportResult(created=created, updated=updated)
else:

    @router.post("/import", status_code=status.HTTP_501_NOT_IMPLEMENTED)
    async def import_products_xlsx_unavailable() -> None:
        raise HTTPException(
            status_code=501,
            detail='Missing dependency: python-multipart. Install with "pip install python-multipart".',
        )


@router.get("/image-files/{filename}")
def get_product_image_file(filename: str) -> FileResponse:
    file_path = _resolve_managed_product_image_path(filename)
    return FileResponse(file_path)


@router.get("/{product_id}", response_model=ProductRead)
def get_product(product_id: int, db: Session = Depends(get_db)) -> Product:
    stmt = select(Product).where(Product.id == product_id).options(selectinload(Product.category))
    product = db.scalar(stmt)
    if product is None:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@router.put("/{product_id}", response_model=ProductRead)
def update_product(product_id: int, product_in: ProductUpdate, db: Session = Depends(get_db)) -> Product:
    product = db.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=404, detail="Product not found")

    data = product_in.model_dump(exclude_unset=True)
    current_sku_upper = (product.sku or "").upper()
    existing_skus_upper = {sku.upper() for sku in db.scalars(select(Product.sku)).all() if sku}
    if "category_id" in data and data["category_id"] is not None:
        cat = db.get(Category, data["category_id"])
        if cat is None:
            raise HTTPException(status_code=400, detail="Invalid category_id")
    if "name" in data and data["name"] is not None:
        data["name"] = data["name"].strip()
    if "currency" in data and data["currency"] is None:
        data["currency"] = settings.DEFAULT_CURRENCY
    if "base_uom" in data and data["base_uom"] is not None:
        data["base_uom"] = data["base_uom"].strip()
    if "base_uom" in data and data["base_uom"] is None:
        data["base_uom"] = "Pc"
    if "uom" in data and data["uom"] is not None:
        data["uom"] = data["uom"].strip()
    if "uom" in data and data["uom"] is None:
        data["uom"] = "Pc"
    if "uom" in data and "uom_multiplier" not in data:
        data["uom_multiplier"] = 12 if (data["uom"] or "").lower() == "dozen" else 1
    if "uom_multiplier" in data and data["uom_multiplier"] is None:
        uom = data.get("uom", product.uom)
        data["uom_multiplier"] = 12 if (uom or "").lower() == "dozen" else 1
    if "uom_multiplier" in data and data["uom_multiplier"] is not None and data["uom_multiplier"] <= 1:
        data["uom_multiplier"] = 1
        data["uom"] = data.get("base_uom", product.base_uom)
    if "sku" in data:
        resolved_name = str(data.get("name", product.name) or "").strip()
        if not resolved_name:
            raise HTTPException(status_code=400, detail="Product name is required")
        data["sku"] = _resolve_sku(
            raw_sku=data["sku"],
            name=resolved_name,
            existing_skus_upper=existing_skus_upper,
            current_sku_upper=current_sku_upper,
        )
    if "is_sold_on_amazon" in data or "amazon_sku" in data:
        is_sold_on_amazon, amazon_sku = _prepare_amazon_fields(
            db,
            is_sold_on_amazon=bool(
                data.get("is_sold_on_amazon", product.is_sold_on_amazon)
            ),
            amazon_sku=data.get("amazon_sku", product.amazon_sku),
            current_product_id=product.id,
        )
        data["is_sold_on_amazon"] = is_sold_on_amazon
        data["amazon_sku"] = amazon_sku
    for key, value in data.items():
        setattr(product, key, value)
    db.flush()
    _sync_product_amazon_mapping(db, product)
    db.commit()
    db.refresh(product)
    return product


@router.patch("/{product_id}", response_model=ProductRead)
def patch_product(product_id: int, product_in: ProductUpdate, db: Session = Depends(get_db)) -> Product:
    # Same semantics as PUT in this MVP: partial update via exclude_unset=True
    return update_product(product_id=product_id, product_in=product_in, db=db)


@router.delete("/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_product(product_id: int, db: Session = Depends(get_db)) -> None:
    product = db.get(Product, product_id)
    if product is None:
        return
    _delete_managed_product_image(product.image_url)
    db.delete(product)
    db.commit()


if _has_python_multipart():

    @router.post("/{product_id}/image", response_model=ProductRead)
    async def upload_product_image(
        product_id: int,
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
    ) -> Product:
        product = db.get(Product, product_id)
        if product is None:
            raise HTTPException(status_code=404, detail="Product not found")

        suffix = _detect_image_suffix(file)
        payload = await file.read()
        if not payload:
            raise HTTPException(status_code=400, detail="Empty image upload")
        if len(payload) > MAX_IMAGE_BYTES:
            raise HTTPException(status_code=400, detail="Image too large. Max size is 10MB.")

        _ensure_product_image_dir()
        filename = f"product-{product_id}-{uuid4().hex}{suffix}"
        file_path = PRODUCT_IMAGE_DIR / filename
        file_path.write_bytes(payload)

        previous_image_url = product.image_url
        product.image_url = f"{PRODUCT_IMAGE_URL_PREFIX}{filename}"
        db.commit()
        db.refresh(product)
        _delete_managed_product_image(previous_image_url)
        return product
else:

    @router.post("/{product_id}/image", status_code=status.HTTP_501_NOT_IMPLEMENTED)
    async def upload_product_image_unavailable(product_id: int) -> None:
        raise HTTPException(
            status_code=501,
            detail='Missing dependency: python-multipart. Install with "pip install python-multipart".',
        )


@router.delete("/{product_id}/image", response_model=ProductRead)
def delete_product_image(product_id: int, db: Session = Depends(get_db)) -> Product:
    product = db.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=404, detail="Product not found")

    previous_image_url = product.image_url
    product.image_url = None
    db.commit()
    db.refresh(product)
    _delete_managed_product_image(previous_image_url)
    return product
