from __future__ import annotations

import csv
from dataclasses import dataclass
from decimal import Decimal
import hashlib
from io import BytesIO, StringIO
import re
from typing import Iterable


BOX_UNITS_RE = re.compile(r"^box\s+(\d+)\s+units$", re.IGNORECASE)


@dataclass(frozen=True)
class AmazonCsvItem:
    row_index: int
    amazon_sku: str
    title: str
    asin: str | None
    fnsku: str | None
    requested_quantity: int


@dataclass(frozen=True)
class AmazonCsvBox:
    number: int
    name: str | None
    weight_lb: float | None
    length_in: float | None
    width_in: float | None
    height_in: float | None


@dataclass
class ParsedAmazonCsv:
    rows: list[list[str]]
    header_row_index: int
    quantity_column_index: int
    item_rows: tuple[AmazonCsvItem, ...]
    box_columns: tuple[tuple[int, int], ...]
    metadata_rows: dict[str, tuple[int, int]]
    pack_group_number: str
    workflow_name: str
    declared_sku_count: int
    declared_unit_count: int
    boxes: tuple[AmazonCsvBox, ...]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class SolverSku:
    amazon_sku: str
    title: str | None
    requested_quantity: int
    available_quantity: int
    unit_weight_lb: float | None
    capacities: dict[int, int]


@dataclass(frozen=True)
class SolverBoxType:
    id: int
    name: str
    length_in: float
    width_in: float
    height_in: float
    empty_weight_lb: float
    max_weight_lb: float | None


def _clean(value: object) -> str:
    return str(value or "").replace("\ufeff", "").strip()


def _parse_positive_int(value: str, *, field: str) -> int:
    raw = _clean(value).replace(",", "")
    try:
        parsed = int(raw)
    except Exception as exc:
        raise ValueError(f"Invalid {field}: {value!r}") from exc
    if parsed < 0:
        raise ValueError(f"Invalid {field}: {value!r}")
    return parsed


def _float_or_none(value: str | None) -> float | None:
    raw = _clean(value)
    if not raw:
        return None
    try:
        return float(raw.replace(",", ""))
    except Exception:
        return None


def _row_value(row: list[str], index: int) -> str:
    return row[index] if 0 <= index < len(row) else ""


def _find_summary_value(rows: list[list[str]], label: str) -> str:
    wanted = label.casefold()
    for row in rows:
        if row and _clean(row[0]).casefold() == wanted:
            return _clean(_row_value(row, 1))
    return ""


def _find_metadata_row(rows: list[list[str]], label: str) -> tuple[int, int] | None:
    wanted = label.casefold()
    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            if _clean(value).casefold() == wanted:
                return row_index, column_index
    return None


def parse_amazon_pack_csv(source: bytes | str) -> ParsedAmazonCsv:
    if isinstance(source, bytes):
        try:
            text = source.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("Amazon packing file must be UTF-8 CSV") from exc
    else:
        text = source.lstrip("\ufeff")

    try:
        rows = [list(row) for row in csv.reader(StringIO(text))]
    except csv.Error as exc:
        raise ValueError(f"Invalid Amazon CSV: {exc}") from exc
    if not rows:
        raise ValueError("Amazon packing CSV is empty")

    header_row_index = -1
    quantity_column_index = -1
    box_columns: list[tuple[int, int]] = []
    for row_index, row in enumerate(rows):
        normalized = [_clean(value).casefold() for value in row]
        if "sku" not in normalized or "quantity" not in normalized:
            continue
        header_row_index = row_index
        quantity_column_index = normalized.index("quantity")
        for column_index, value in enumerate(row):
            matched = BOX_UNITS_RE.match(_clean(value))
            if matched:
                box_columns.append((int(matched.group(1)), column_index))
        break
    if header_row_index < 0 or quantity_column_index < 0:
        raise ValueError("Could not find the SKU/Quantity header in the Amazon CSV")
    box_columns.sort(key=lambda entry: entry[0])
    if not box_columns:
        raise ValueError("Could not find any 'Box N units' columns in the Amazon CSV")

    header = rows[header_row_index]
    header_lookup = {_clean(value).casefold(): index for index, value in enumerate(header)}
    sku_index = header_lookup["sku"]
    title_index = header_lookup.get("title", 1)
    asin_index = header_lookup.get("asin", 2)
    fnsku_index = header_lookup.get("fnsku", 3)

    items: list[AmazonCsvItem] = []
    for row_index in range(header_row_index + 1, len(rows)):
        row = rows[row_index]
        sku = _clean(_row_value(row, sku_index))
        if not sku:
            if items:
                break
            continue
        quantity = _parse_positive_int(
            _row_value(row, quantity_column_index),
            field=f"quantity for SKU {sku}",
        )
        if quantity < 1:
            raise ValueError(f"Quantity for SKU {sku} must be greater than zero")
        items.append(
            AmazonCsvItem(
                row_index=row_index,
                amazon_sku=sku,
                title=_clean(_row_value(row, title_index)),
                asin=_clean(_row_value(row, asin_index)) or None,
                fnsku=_clean(_row_value(row, fnsku_index)) or None,
                requested_quantity=quantity,
            )
        )
    if not items:
        raise ValueError("Amazon packing CSV does not contain any SKU rows")

    metadata_labels = {
        "box_id": "Box ID",
        "box_name": "Box name",
        "weight": "Box weight (lb):",
        "length": "Box length (inch):",
        "width": "Box width (inch):",
        "height": "Box height (inch):",
    }
    metadata_rows: dict[str, tuple[int, int]] = {}
    for key, label in metadata_labels.items():
        found = _find_metadata_row(rows, label)
        if found is None:
            raise ValueError(f"Could not find '{label}' in the Amazon CSV")
        metadata_rows[key] = found

    metadata_box_count = max(
        max(0, len(rows[row_index]) - label_column_index - 1)
        for row_index, label_column_index in metadata_rows.values()
    )
    existing_box_count = max(len(box_columns), metadata_box_count)

    def metadata_values(key: str) -> list[str]:
        row_index, label_column_index = metadata_rows[key]
        row = rows[row_index]
        return [_row_value(row, label_column_index + 1 + index) for index in range(existing_box_count)]

    names = metadata_values("box_name")
    weights = metadata_values("weight")
    lengths = metadata_values("length")
    widths = metadata_values("width")
    heights = metadata_values("height")
    boxes = tuple(
        AmazonCsvBox(
            number=index + 1,
            name=_clean(names[index]) or None,
            weight_lb=_float_or_none(weights[index]),
            length_in=_float_or_none(lengths[index]),
            width_in=_float_or_none(widths[index]),
            height_in=_float_or_none(heights[index]),
        )
        for index in range(existing_box_count)
    )

    declared_sku_raw = _find_summary_value(rows, "SKUs")
    declared_units_raw = _find_summary_value(rows, "Units")
    declared_sku_count = _parse_positive_int(declared_sku_raw or str(len(items)), field="SKU count")
    declared_unit_count = _parse_positive_int(
        declared_units_raw or str(sum(item.requested_quantity for item in items)),
        field="unit count",
    )
    warnings: list[str] = []
    if declared_sku_count != len(items):
        warnings.append(
            f"CSV declares {declared_sku_count} SKUs but contains {len(items)} SKU rows."
        )
    actual_units = sum(item.requested_quantity for item in items)
    if declared_unit_count != actual_units:
        warnings.append(
            f"CSV declares {declared_unit_count} units but SKU rows total {actual_units}."
        )

    return ParsedAmazonCsv(
        rows=rows,
        header_row_index=header_row_index,
        quantity_column_index=quantity_column_index,
        item_rows=tuple(items),
        box_columns=tuple(box_columns),
        metadata_rows=metadata_rows,
        pack_group_number=_find_summary_value(rows, "Pack group number"),
        workflow_name=_find_summary_value(rows, "Workflow name"),
        declared_sku_count=declared_sku_count,
        declared_unit_count=declared_unit_count,
        boxes=boxes,
        warnings=tuple(warnings),
    )


def _decimal_text(value: Decimal | float) -> str:
    decimal_value = value if isinstance(value, Decimal) else Decimal(str(value))
    normalized = format(decimal_value.normalize(), "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    return normalized or "0"


def render_amazon_pack_csv(
    source: bytes | str,
    *,
    per_box_quantities: dict[str, int],
    boxes: list[dict[str, Decimal | str]],
) -> bytes:
    parsed = parse_amazon_pack_csv(source)
    if len(boxes) < 5:
        raise ValueError("Amazon-optimized export requires at least five boxes")

    source_skus = {item.amazon_sku for item in parsed.item_rows}
    plan_skus = set(per_box_quantities)
    if source_skus != plan_skus:
        missing = sorted(source_skus - plan_skus)
        extra = sorted(plan_skus - source_skus)
        details: list[str] = []
        if missing:
            details.append(f"missing SKUs: {', '.join(missing)}")
        if extra:
            details.append(f"unknown SKUs: {', '.join(extra)}")
        raise ValueError("Export plan does not match the Amazon template (" + "; ".join(details) + ")")
    for sku, quantity in per_box_quantities.items():
        if quantity < 1:
            raise ValueError(f"Per-box quantity for {sku} must be greater than zero")

    rows = [list(row) for row in parsed.rows]
    box_count = len(boxes)
    total_units = sum(per_box_quantities[item.amazon_sku] * box_count for item in parsed.item_rows)

    for row in rows:
        if row and _clean(row[0]).casefold() == "skus":
            while len(row) < 2:
                row.append("")
            row[1] = str(len(parsed.item_rows))
        if row and _clean(row[0]).casefold() == "units":
            while len(row) < 2:
                row.append("")
            row[1] = str(total_units)

    header = rows[parsed.header_row_index]
    rows[parsed.header_row_index] = header[: parsed.quantity_column_index + 1] + [
        f"Box {index + 1} units" for index in range(box_count)
    ]

    for item in parsed.item_rows:
        row = rows[item.row_index]
        prefix = row[: parsed.quantity_column_index + 1]
        while len(prefix) <= parsed.quantity_column_index:
            prefix.append("")
        per_box = per_box_quantities[item.amazon_sku]
        prefix[parsed.quantity_column_index] = str(per_box * box_count)
        rows[item.row_index] = prefix + [str(per_box)] * box_count

    metadata_values: dict[str, list[str]] = {
        "box_id": ["To be assigned"] * box_count,
        "box_name": [str(box["name"]) for box in boxes],
        "weight": [_decimal_text(box["weight_lb"]) for box in boxes],
        "length": [_decimal_text(box["length_in"]) for box in boxes],
        "width": [_decimal_text(box["width_in"]) for box in boxes],
        "height": [_decimal_text(box["height_in"]) for box in boxes],
    }
    for key, values in metadata_values.items():
        row_index, label_column_index = parsed.metadata_rows[key]
        row = rows[row_index]
        rows[row_index] = row[: label_column_index + 1] + values

    output = StringIO(newline="")
    writer = csv.writer(output, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
    writer.writerows(rows)
    result = ("\ufeff" + output.getvalue()).encode("utf-8")

    verified = parse_amazon_pack_csv(result)
    if len(verified.box_columns) != box_count:
        raise RuntimeError("Generated Amazon CSV has an invalid box column count")
    if verified.declared_unit_count != total_units:
        raise RuntimeError("Generated Amazon CSV has an invalid unit total")
    return result


def render_amazon_manifest_xlsx(
    source: bytes,
    *,
    items: list[tuple[str, int]],
) -> bytes:
    if not items:
        raise ValueError("Select at least one Amazon SKU")
    seen_skus: set[str] = set()
    for amazon_sku, quantity in items:
        if amazon_sku in seen_skus:
            raise ValueError(f"Amazon SKU appears more than once: {amazon_sku}")
        seen_skus.add(amazon_sku)
        if not amazon_sku or len(amazon_sku) > 80 or not amazon_sku.isascii():
            raise ValueError(f"Invalid Amazon SKU: {amazon_sku!r}")
        if quantity < 1:
            raise ValueError(f"Quantity for {amazon_sku} must be greater than zero")

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("Missing dependency: openpyxl") from exc
    try:
        workbook = load_workbook(BytesIO(source), data_only=False, keep_links=True)
    except Exception as exc:
        raise ValueError("Could not open the Amazon manifest XLSX template") from exc

    template_sheet = next(
        (
            sheet
            for sheet in workbook.worksheets
            if sheet.title.strip().casefold() == "create workflow – template".casefold()
        ),
        None,
    )
    header_row = 0
    sku_column = 0
    quantity_column = 0
    candidate_sheets = [template_sheet] if template_sheet is not None else list(workbook.worksheets)
    for sheet in candidate_sheets:
        if sheet is None:
            continue
        for row_index in range(1, min(sheet.max_row, 40) + 1):
            normalized = {
                str(sheet.cell(row_index, column_index).value or "").strip().casefold(): column_index
                for column_index in range(1, min(sheet.max_column, 40) + 1)
            }
            if "merchant sku" in normalized and "quantity" in normalized:
                template_sheet = sheet
                header_row = row_index
                sku_column = normalized["merchant sku"]
                quantity_column = normalized["quantity"]
                break
        if header_row:
            break
    if template_sheet is None or not header_row:
        raise ValueError(
            "Could not find the Merchant SKU and Quantity columns in the Amazon template"
        )

    for row_index in range(header_row + 1, template_sheet.max_row + 1):
        for column_index in range(1, template_sheet.max_column + 1):
            template_sheet.cell(row_index, column_index).value = None
    for offset, (amazon_sku, quantity) in enumerate(items, start=1):
        row_index = header_row + offset
        template_sheet.cell(row_index, sku_column).value = amazon_sku
        template_sheet.cell(row_index, quantity_column).value = quantity

    workbook.active = workbook.worksheets.index(template_sheet)
    output = BytesIO()
    workbook.save(output)
    result = output.getvalue()

    try:
        verified = load_workbook(BytesIO(result), read_only=True, data_only=False)
        verified_sheet = verified[template_sheet.title]
        for offset, (amazon_sku, quantity) in enumerate(items, start=1):
            row_index = header_row + offset
            if verified_sheet.cell(row_index, sku_column).value != amazon_sku:
                raise RuntimeError("Generated manifest has an invalid Amazon SKU row")
            if verified_sheet.cell(row_index, quantity_column).value != quantity:
                raise RuntimeError("Generated manifest has an invalid quantity row")
    except RuntimeError:
        raise
    except Exception as exc:
        raise RuntimeError("Generated Amazon manifest could not be verified") from exc
    return result


def render_amazon_pack_xlsx(
    source: bytes,
    *,
    per_box_quantities: dict[str, int],
    boxes: list[dict[str, object]],
) -> bytes:
    if not per_box_quantities:
        raise ValueError("The optimized plan does not contain any Amazon SKUs")
    if len(boxes) < 5:
        raise ValueError("Amazon optimized shipments require at least 5 boxes")
    if any(quantity < 1 for quantity in per_box_quantities.values()):
        raise ValueError("Every SKU must have a positive per-box quantity")

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("Missing dependency: openpyxl") from exc
    try:
        workbook = load_workbook(BytesIO(source), data_only=False, keep_links=True)
    except Exception as exc:
        raise ValueError("Could not open the Amazon Box Packing Information XLSX file") from exc

    target_sheet = None
    header_row = 0
    sku_column = 0
    expected_column = 0
    first_box_column = 0
    for sheet in workbook.worksheets:
        for row_index in range(1, min(sheet.max_row, 60) + 1):
            normalized = {
                _clean(sheet.cell(row_index, column_index).value).casefold(): column_index
                for column_index in range(1, min(sheet.max_column, 80) + 1)
            }
            if not {"sku", "expected quantity", "boxed quantity"}.issubset(normalized):
                continue
            for column_index in range(normalized["boxed quantity"] + 1, sheet.max_column + 1):
                value = _clean(sheet.cell(row_index, column_index).value)
                if re.search(r"box\s+1\s+quantity", value, re.IGNORECASE):
                    first_box_column = column_index
                    break
            if first_box_column:
                target_sheet = sheet
                header_row = row_index
                sku_column = normalized["sku"]
                expected_column = normalized["expected quantity"]
                break
        if target_sheet is not None:
            break
    if target_sheet is None or not first_box_column:
        raise ValueError(
            "Could not find the SKU and box quantity columns in the Amazon XLSX file"
        )

    box_columns = list(range(first_box_column, target_sheet.max_column + 1))
    if len(boxes) > len(box_columns):
        raise ValueError(
            f"This Amazon XLSX supports at most {len(box_columns)} boxes, but the plan uses {len(boxes)}"
        )

    sku_rows: dict[str, int] = {}
    for row_index in range(header_row + 1, target_sheet.max_row + 1):
        sku = _clean(target_sheet.cell(row_index, sku_column).value)
        if not sku:
            if sku_rows:
                break
            continue
        if sku in sku_rows:
            raise ValueError(f"Amazon XLSX contains duplicate SKU rows: {sku}")
        sku_rows[sku] = row_index
    if not sku_rows:
        raise ValueError("Amazon XLSX does not contain any SKU rows")

    missing_skus = sorted(set(per_box_quantities) - set(sku_rows))
    extra_skus = sorted(set(sku_rows) - set(per_box_quantities))
    if missing_skus or extra_skus:
        details: list[str] = []
        if missing_skus:
            details.append("missing from XLSX: " + ", ".join(missing_skus))
        if extra_skus:
            details.append("not in selected plan: " + ", ".join(extra_skus))
        raise ValueError("Amazon XLSX SKU list does not match the optimized plan (" + "; ".join(details) + ")")

    box_count = len(boxes)
    for amazon_sku, row_index in sku_rows.items():
        per_box_quantity = per_box_quantities[amazon_sku]
        expected_quantity = target_sheet.cell(row_index, expected_column).value
        try:
            expected_int = int(expected_quantity)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid expected quantity for {amazon_sku}") from exc
        optimized_total = per_box_quantity * box_count
        if expected_int != optimized_total:
            raise ValueError(
                f"Amazon expects {expected_int} units for {amazon_sku}, but the selected plan uses "
                f"{optimized_total}. Update the workflow quantity before filling this XLSX."
            )
        for offset, column_index in enumerate(box_columns):
            target_sheet.cell(row_index, column_index).value = (
                per_box_quantity if offset < box_count else None
            )

    total_box_label_row = 0
    detail_rows: dict[str, int] = {}
    detail_labels = {
        "name": "name of box",
        "weight_lb": "box weight (lb):",
        "width_in": "box width (inch):",
        "length_in": "box length (inch):",
        "height_in": "box height (inch):",
    }
    for row_index in range(1, target_sheet.max_row + 1):
        for column_index in range(1, min(first_box_column, target_sheet.max_column + 1)):
            value = _clean(target_sheet.cell(row_index, column_index).value).casefold()
            if value == "total box count:":
                total_box_label_row = row_index
            for key, label in detail_labels.items():
                if value == label:
                    detail_rows[key] = row_index
    if not total_box_label_row or set(detail_rows) != set(detail_labels):
        raise ValueError("Could not find the box count, weight, and dimension rows in the Amazon XLSX")

    target_sheet.cell(total_box_label_row, first_box_column).value = box_count
    for offset, column_index in enumerate(box_columns):
        if offset < box_count:
            box = boxes[offset]
            target_sheet.cell(detail_rows["name"], column_index).value = str(box["name"])
            target_sheet.cell(detail_rows["weight_lb"], column_index).value = float(box["weight_lb"])
            target_sheet.cell(detail_rows["width_in"], column_index).value = float(box["width_in"])
            target_sheet.cell(detail_rows["length_in"], column_index).value = float(box["length_in"])
            target_sheet.cell(detail_rows["height_in"], column_index).value = float(box["height_in"])
        else:
            for key in ("weight_lb", "width_in", "length_in", "height_in"):
                target_sheet.cell(detail_rows[key], column_index).value = None

    target_sheet.sheet_view.showGridLines = False
    workbook.active = workbook.worksheets.index(target_sheet)
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except Exception:
        pass

    output = BytesIO()
    workbook.save(output)
    result = output.getvalue()

    try:
        verified = load_workbook(BytesIO(result), read_only=True, data_only=False)
        verified_sheet = verified[target_sheet.title]
        if verified_sheet.cell(total_box_label_row, first_box_column).value != box_count:
            raise RuntimeError("Generated Amazon XLSX has an invalid box count")
        for amazon_sku, row_index in sku_rows.items():
            for offset in range(box_count):
                if verified_sheet.cell(row_index, first_box_column + offset).value != per_box_quantities[amazon_sku]:
                    raise RuntimeError("Generated Amazon XLSX has invalid box quantities")
    except RuntimeError:
        raise
    except Exception as exc:
        raise RuntimeError("Generated Amazon Box Packing XLSX could not be verified") from exc
    return result


def _capacity_utilization(content: list[int], skus: list[SolverSku], box_type_id: int) -> float:
    return sum(
        quantity / skus[index].capacities[box_type_id]
        for index, quantity in enumerate(content)
    )


def _estimated_weight(content: list[int], skus: list[SolverSku], box: SolverBoxType) -> float | None:
    if any(sku.unit_weight_lb is None for sku in skus):
        return None
    return box.empty_weight_lb + sum(
        content[index] * float(sku.unit_weight_lb or 0)
        for index, sku in enumerate(skus)
    )


def _fits(content: list[int], skus: list[SolverSku], box: SolverBoxType) -> bool:
    if _capacity_utilization(content, skus, box.id) > 1.0 + 1e-9:
        return False
    estimated_weight = _estimated_weight(content, skus, box)
    if estimated_weight is not None and box.max_weight_lb is not None:
        return estimated_weight <= box.max_weight_lb + 1e-9
    return True


def _quantity_error(content: list[int], skus: list[SolverSku], box_count: int) -> int:
    return sum(
        abs(box_count * content[index] - sku.requested_quantity)
        for index, sku in enumerate(skus)
    )


def _violation(content: list[int], skus: list[SolverSku], box: SolverBoxType) -> float:
    capacity_over = max(0.0, _capacity_utilization(content, skus, box.id) - 1.0)
    weight = _estimated_weight(content, skus, box)
    weight_over = 0.0
    if weight is not None and box.max_weight_lb:
        weight_over = max(0.0, (weight - box.max_weight_lb) / box.max_weight_lb)
    return capacity_over + weight_over


def _nearest_content(skus: list[SolverSku], box: SolverBoxType, box_count: int) -> list[int] | None:
    content: list[int] = []
    for sku in skus:
        capacity = sku.capacities.get(box.id)
        if not capacity:
            return None
        max_per_box = min(capacity, sku.available_quantity // box_count)
        if max_per_box < 1:
            return None
        best = min(
            range(1, max_per_box + 1),
            key=lambda quantity: (
                abs(box_count * quantity - sku.requested_quantity),
                -box_count * quantity,
            ),
        )
        content.append(best)

    guard = sum(content) + len(content) + 1
    while not _fits(content, skus, box) and guard > 0:
        guard -= 1
        current_violation = _violation(content, skus, box)
        current_error = _quantity_error(content, skus, box_count)
        options: list[tuple[float, float, int]] = []
        for index, quantity in enumerate(content):
            if quantity <= 1:
                continue
            candidate = list(content)
            candidate[index] -= 1
            improvement = current_violation - _violation(candidate, skus, box)
            if improvement <= 1e-12:
                continue
            error_increase = _quantity_error(candidate, skus, box_count) - current_error
            options.append((error_increase / improvement, -improvement, index))
        if not options:
            return None
        _, _, selected_index = min(options)
        content[selected_index] -= 1
    if not _fits(content, skus, box):
        return None

    # A capacity-driven decrement can move an item farther below its requested
    # quantity than necessary. Add units back only when doing so reduces the
    # total quantity adjustment and still fits.
    while True:
        current_error = _quantity_error(content, skus, box_count)
        improvements: list[tuple[int, float, int]] = []
        for index, sku in enumerate(skus):
            max_per_box = min(sku.capacities[box.id], sku.available_quantity // box_count)
            if content[index] >= max_per_box:
                continue
            candidate = list(content)
            candidate[index] += 1
            if not _fits(candidate, skus, box):
                continue
            error_reduction = current_error - _quantity_error(candidate, skus, box_count)
            if error_reduction > 0:
                improvements.append(
                    (-error_reduction, -1 / sku.capacities[box.id], index)
                )
        if not improvements:
            break
        _, _, selected_index = min(improvements)
        content[selected_index] += 1
    return content


def _max_fill_content(
    balanced: list[int],
    skus: list[SolverSku],
    box: SolverBoxType,
    box_count: int,
) -> list[int]:
    content = list(balanced)
    guard = sum(sku.capacities[box.id] for sku in skus) + 1
    while guard > 0:
        guard -= 1
        options: list[tuple[float, int, int]] = []
        current_error = _quantity_error(content, skus, box_count)
        for index, sku in enumerate(skus):
            max_per_box = min(sku.capacities[box.id], sku.available_quantity // box_count)
            if content[index] >= max_per_box:
                continue
            candidate = list(content)
            candidate[index] += 1
            if not _fits(candidate, skus, box):
                continue
            capacity_step = 1 / sku.capacities[box.id]
            error_increase = _quantity_error(candidate, skus, box_count) - current_error
            options.append((error_increase, -capacity_step, index))
        if not options:
            break
        _, _, selected_index = min(options)
        content[selected_index] += 1
    return content


def _feasible_boxes(
    content: list[int],
    skus: list[SolverSku],
    box_types: list[SolverBoxType],
) -> list[dict[str, float | int | str | None]]:
    feasible: list[dict[str, float | int | str | None]] = []
    for box in box_types:
        if any(box.id not in sku.capacities for sku in skus):
            continue
        if not _fits(content, skus, box):
            continue
        feasible.append(
            {
                "id": box.id,
                "name": box.name,
                "length_in": box.length_in,
                "width_in": box.width_in,
                "height_in": box.height_in,
                "empty_weight_lb": box.empty_weight_lb,
                "max_weight_lb": box.max_weight_lb,
                "capacity_utilization": _capacity_utilization(content, skus, box.id),
                "estimated_weight_lb": _estimated_weight(content, skus, box),
            }
        )
    feasible.sort(key=lambda value: (-float(value["capacity_utilization"] or 0), str(value["name"])))
    return feasible


def _plan_dict(
    *,
    strategy: str,
    content: list[int],
    skus: list[SolverSku],
    box_count: int,
    box_types: list[SolverBoxType],
) -> dict[str, object] | None:
    feasible = _feasible_boxes(content, skus, box_types)
    if not feasible:
        return None
    selected = feasible[0]
    selected_box_id = int(selected["id"])
    signature = f"{box_count}|" + "|".join(
        f"{sku.amazon_sku}:{content[index]}" for index, sku in enumerate(skus)
    )
    warnings: list[str] = []
    if any(sku.unit_weight_lb is None for sku in skus):
        warnings.append("Estimated weight is unavailable until every Amazon SKU has a unit weight.")
    items: list[dict[str, object]] = []
    for index, sku in enumerate(skus):
        per_box = content[index]
        adjusted = per_box * box_count
        capacity_units = sku.capacities[selected_box_id]
        items.append(
            {
                "amazon_sku": sku.amazon_sku,
                "title": sku.title,
                "requested_quantity": sku.requested_quantity,
                "available_quantity": sku.available_quantity,
                "per_box_quantity": per_box,
                "adjusted_quantity": adjusted,
                "quantity_delta": adjusted - sku.requested_quantity,
                "unit_weight_lb": sku.unit_weight_lb,
                "capacity_units": capacity_units,
                "capacity_fraction": per_box / capacity_units,
            }
        )
    requested_total = sum(sku.requested_quantity for sku in skus)
    adjusted_total = sum(item["adjusted_quantity"] for item in items)
    absolute_change = sum(abs(int(item["quantity_delta"])) for item in items)
    if absolute_change:
        warnings.append(
            "Update the SKU quantities in the Amazon workflow so they match this plan before downloading the Box Packing XLSX."
        )
    return {
        "key": hashlib.sha1(signature.encode("utf-8")).hexdigest()[:16],
        "strategy": strategy,
        "box_count": box_count,
        "selected_box_type_id": selected_box_id,
        "selected_box_type_name": str(selected["name"]),
        "requested_unit_count": requested_total,
        "adjusted_unit_count": adjusted_total,
        "absolute_quantity_change": absolute_change,
        "units_per_box": sum(content),
        "capacity_utilization": float(selected["capacity_utilization"] or 0),
        "estimated_weight_lb": selected["estimated_weight_lb"],
        "items": items,
        "feasible_box_types": feasible,
        "warnings": warnings,
    }


def optimize_identical_cartons(
    *,
    skus: Iterable[SolverSku],
    box_types: Iterable[SolverBoxType],
    min_box_count: int,
    max_box_count: int,
) -> list[dict[str, object]]:
    sku_list = list(skus)
    box_list = list(box_types)
    if not sku_list or not box_list:
        return []

    candidates: dict[tuple[int, tuple[int, ...]], dict[str, object]] = {}
    max_possible_boxes = min(
        max_box_count,
        min(sku.available_quantity for sku in sku_list),
    )
    for box_count in range(min_box_count, max_possible_boxes + 1):
        for box in box_list:
            if any(box.id not in sku.capacities for sku in sku_list):
                continue
            balanced = _nearest_content(sku_list, box, box_count)
            if balanced is None:
                continue
            for strategy, content in (
                ("Closest to requested list", balanced),
                ("Maximum box fill", _max_fill_content(balanced, sku_list, box, box_count)),
            ):
                signature = (box_count, tuple(content))
                if signature in candidates:
                    continue
                plan = _plan_dict(
                    strategy=strategy,
                    content=content,
                    skus=sku_list,
                    box_count=box_count,
                    box_types=box_list,
                )
                if plan is not None:
                    candidates[signature] = plan

    plans = list(candidates.values())
    closest = sorted(
        (plan for plan in plans if plan["strategy"] == "Closest to requested list"),
        key=lambda plan: (
            int(plan["absolute_quantity_change"]),
            -float(plan["capacity_utilization"]),
            int(plan["box_count"]),
            -int(plan["adjusted_unit_count"]),
        ),
    )[:4]
    max_fill = sorted(
        (plan for plan in plans if plan["strategy"] == "Maximum box fill"),
        key=lambda plan: (
            -float(plan["capacity_utilization"]),
            int(plan["absolute_quantity_change"]),
            int(plan["box_count"]),
        ),
    )[:3]
    selected: list[dict[str, object]] = []
    seen_keys: set[str] = set()
    for plan in [*closest, *max_fill]:
        key = str(plan["key"])
        if key in seen_keys:
            continue
        selected.append(plan)
        seen_keys.add(key)
    return selected
