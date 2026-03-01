from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from app.db.models import Product


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _openpyxl():
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore
        from openpyxl.styles import Font  # type: ignore
    except Exception as e:  # pragma: no cover
        raise RuntimeError("Missing dependency: openpyxl. Please `pip install openpyxl`.") from e
    return Workbook, load_workbook, Font


def _wb_bytes(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_style(ws) -> None:
    _, _, Font = _openpyxl()
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _cell_dt(v: Any) -> datetime | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    # Accept ISO-ish strings
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        return None


def build_receipt_template_xlsx() -> bytes:
    Workbook, _, _ = _openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipt Import"
    ws.append(
        [
            "receipt_number",
            "received_at",
            "received_by",
            "note",
            "sku",
            "quantity",
            "unit",  # BASE or SALE
            "unit_cost",
            "line_note",
        ]
    )
    ws.append(
        [
            "R-0001",
            "",  # can be Excel datetime
            "",
            "",
            "UL1628T001",
            10,
            "BASE",
            0,
            "",
        ]
    )
    _header_style(ws)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 18
    return _wb_bytes(wb)


def build_issue_template_xlsx() -> bytes:
    Workbook, _, _ = _openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Issue Import"
    ws.append(
        [
            "issue_number",
            "issued_at",
            "issued_by",
            "issued_to",
            "purpose",
            "note",
            "sku",
            "product_id",
            "quantity",
            "unit",  # BASE or SALE
            "line_note",
        ]
    )
    ws.append(["I-0001", "", "", "", "TEST", "", "UL1628T001", "", 1, "BASE", ""])
    _header_style(ws)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 18
    return _wb_bytes(wb)


class ExcelImportError(ValueError):
    def __init__(self, errors: list[str]):
        super().__init__("\n".join(errors))
        self.errors = errors


def _read_rows_from_xlsx(file_bytes: bytes) -> tuple[list[str], list[list[Any]]]:
    _, load_workbook, _ = _openpyxl()
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ExcelImportError(["Empty workbook"])
    header = [(_cell_str(x) or "").strip() for x in (rows[0] or [])]
    header_lc = [h.lower() for h in header]
    data_rows: list[list[Any]] = [list(r or []) for r in rows[1:]]
    return header_lc, data_rows


def parse_receipt_import_xlsx(file_bytes: bytes, *, products_by_sku: dict[str, Product]) -> dict[str, Any]:
    header, rows = _read_rows_from_xlsx(file_bytes)
    errors: list[str] = []
    if "quantity" not in header:
        errors.append("Missing column: quantity")
    if "sku" not in header and "product_id" not in header:
        errors.append("Missing column: sku (or product_id)")
    if errors:
        raise ExcelImportError(errors)

    idx = {name: header.index(name) for name in header}

    receipt_number: str | None = None
    received_at: datetime | None = None
    received_by: str | None = None
    note: str | None = None

    lines: list[dict[str, Any]] = []
    for row_i, row in enumerate(rows, start=2):
        sku = _cell_str(row[idx["sku"]] if "sku" in idx and idx["sku"] < len(row) else "")
        product_id_raw = _cell_str(row[idx["product_id"]] if "product_id" in idx and idx["product_id"] < len(row) else "")
        if not sku and not product_id_raw:
            # stop if fully blank-ish
            continue

        rn = _cell_str(row[idx["receipt_number"]] if "receipt_number" in idx and idx["receipt_number"] < len(row) else "")
        if rn:
            if receipt_number is None:
                receipt_number = rn
            elif rn != receipt_number:
                errors.append(f"Row {row_i}: receipt_number mismatch ({rn} != {receipt_number})")

        dt = _cell_dt(row[idx["received_at"]] if "received_at" in idx and idx["received_at"] < len(row) else None)
        if dt:
            if received_at is None:
                received_at = dt
            elif dt != received_at:
                errors.append(f"Row {row_i}: received_at mismatch")

        rb = _cell_str(row[idx["received_by"]] if "received_by" in idx and idx["received_by"] < len(row) else "")
        if rb:
            if received_by is None:
                received_by = rb

        nt = _cell_str(row[idx["note"]] if "note" in idx and idx["note"] < len(row) else "")
        if nt:
            if note is None:
                note = nt

        qty_raw = row[idx["quantity"]] if "quantity" in idx and idx["quantity"] < len(row) else None
        try:
            qty = int(qty_raw)
        except Exception:
            errors.append(f"Row {row_i}: invalid quantity")
            continue
        if qty <= 0:
            errors.append(f"Row {row_i}: quantity must be > 0")
            continue

        unit = _cell_str(row[idx["unit"]] if "unit" in idx and idx["unit"] < len(row) else "") or "BASE"
        unit = unit.upper()
        if unit not in {"BASE", "SALE"}:
            errors.append(f"Row {row_i}: unit must be BASE or SALE")
            continue

        unit_cost_raw = row[idx["unit_cost"]] if "unit_cost" in idx and idx["unit_cost"] < len(row) else 0
        try:
            unit_cost = float(unit_cost_raw or 0)
        except Exception:
            errors.append(f"Row {row_i}: invalid unit_cost")
            continue
        if unit_cost < 0:
            errors.append(f"Row {row_i}: unit_cost must be >= 0")
            continue

        line_note = _cell_str(row[idx["line_note"]] if "line_note" in idx and idx["line_note"] < len(row) else "")
        if not line_note and nt:
            line_note = nt

        product_id_from_sku: int | None = None
        product_id_from_col: int | None = None

        if sku:
            key = sku.strip().upper()
            product = products_by_sku.get(key)
            if product is None and not product_id_raw:
                errors.append(f"Row {row_i}: unknown SKU {sku!r}")
                continue
            if product is not None:
                product_id_from_sku = int(product.id)

        if product_id_raw:
            try:
                product_id_from_col = int(float(product_id_raw))
            except Exception:
                errors.append(f"Row {row_i}: invalid product_id")
                continue

        if product_id_from_sku is not None and product_id_from_col is not None and product_id_from_sku != product_id_from_col:
            errors.append(
                f"Row {row_i}: sku/product_id mismatch ({product_id_from_sku} != {product_id_from_col})"
            )
            continue

        product_id = product_id_from_sku if product_id_from_sku is not None else product_id_from_col
        if product_id is None:
            errors.append(f"Row {row_i}: missing sku/product_id")
            continue

        lines.append(
            {
                "product_id": product_id,
                "quantity": qty,
                "unit": unit,
                "unit_cost": unit_cost,
                "note": line_note or None,
            }
        )

    if not lines:
        errors.append("No line items found")

    if errors:
        raise ExcelImportError(errors)

    return {
        "receipt_number": receipt_number or None,
        "received_at": received_at,
        "received_by": received_by or None,
        "note": note or None,
        "lines": lines,
    }


def parse_issue_import_xlsx(file_bytes: bytes, *, products_by_sku: dict[str, Product]) -> dict[str, Any]:
    header, rows = _read_rows_from_xlsx(file_bytes)
    errors: list[str] = []
    if "quantity" not in header:
        errors.append("Missing column: quantity")
    if "sku" not in header and "product_id" not in header:
        errors.append("Missing column: sku (or product_id)")
    if errors:
        raise ExcelImportError(errors)

    idx = {name: header.index(name) for name in header}

    issue_number: str | None = None
    issued_at: datetime | None = None
    issued_by: str | None = None
    issued_to: str | None = None
    purpose: str | None = None
    note: str | None = None

    lines: list[dict[str, Any]] = []
    for row_i, row in enumerate(rows, start=2):
        sku = _cell_str(row[idx["sku"]] if "sku" in idx and idx["sku"] < len(row) else "")
        product_id_raw = _cell_str(row[idx["product_id"]] if "product_id" in idx and idx["product_id"] < len(row) else "")
        if not sku and not product_id_raw:
            continue

        inum = _cell_str(row[idx["issue_number"]] if "issue_number" in idx and idx["issue_number"] < len(row) else "")
        if inum:
            if issue_number is None:
                issue_number = inum
            elif inum != issue_number:
                errors.append(f"Row {row_i}: issue_number mismatch ({inum} != {issue_number})")

        dt = _cell_dt(row[idx["issued_at"]] if "issued_at" in idx and idx["issued_at"] < len(row) else None)
        if dt:
            if issued_at is None:
                issued_at = dt
            elif dt != issued_at:
                errors.append(f"Row {row_i}: issued_at mismatch")

        ib = _cell_str(row[idx["issued_by"]] if "issued_by" in idx and idx["issued_by"] < len(row) else "")
        if ib:
            if issued_by is None:
                issued_by = ib
            elif ib != issued_by:
                errors.append(f"Row {row_i}: issued_by mismatch")

        it = _cell_str(row[idx["issued_to"]] if "issued_to" in idx and idx["issued_to"] < len(row) else "")
        if it:
            if issued_to is None:
                issued_to = it
            elif it != issued_to:
                errors.append(f"Row {row_i}: issued_to mismatch")

        pur = _cell_str(row[idx["purpose"]] if "purpose" in idx and idx["purpose"] < len(row) else "")
        if pur:
            pur = pur.strip().upper()
            if purpose is None:
                purpose = pur
            elif pur != purpose:
                errors.append(f"Row {row_i}: purpose mismatch")

        nt = _cell_str(row[idx["note"]] if "note" in idx and idx["note"] < len(row) else "")
        if nt:
            if note is None:
                note = nt
            elif nt != note:
                errors.append(f"Row {row_i}: note mismatch")

        qty_raw = row[idx["quantity"]] if "quantity" in idx and idx["quantity"] < len(row) else None
        try:
            qty = int(qty_raw)
        except Exception:
            errors.append(f"Row {row_i}: invalid quantity")
            continue
        if qty <= 0:
            errors.append(f"Row {row_i}: quantity must be > 0")
            continue

        unit = _cell_str(row[idx["unit"]] if "unit" in idx and idx["unit"] < len(row) else "") or "BASE"
        unit = unit.upper()
        if unit not in {"BASE", "SALE"}:
            errors.append(f"Row {row_i}: unit must be BASE or SALE")
            continue

        line_note = _cell_str(row[idx["line_note"]] if "line_note" in idx and idx["line_note"] < len(row) else "")

        if product_id_raw:
            try:
                pid = int(float(product_id_raw))
            except Exception:
                errors.append(f"Row {row_i}: invalid product_id")
                continue
            product_id = pid
        else:
            key = sku.strip().upper()
            product = products_by_sku.get(key)
            if product is None:
                errors.append(f"Row {row_i}: unknown SKU {sku!r}")
                continue
            product_id = product.id

        lines.append(
            {
                "product_id": product_id,
                "quantity": qty,
                "unit": unit,
                "note": line_note or None,
            }
        )

    if not lines:
        errors.append("No line items found")

    if errors:
        raise ExcelImportError(errors)

    return {
        "issue_number": issue_number or None,
        "issued_at": issued_at,
        "issued_by": issued_by or None,
        "issued_to": issued_to or None,
        "purpose": (purpose or "OTHER"),
        "note": note or None,
        "lines": lines,
    }
