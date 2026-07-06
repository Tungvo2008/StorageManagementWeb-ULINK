from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from app.db.models import InvoiceLineType, Product


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _openpyxl():
    try:
        from openpyxl import Workbook, load_workbook  # type: ignore
        from openpyxl.styles import Font  # type: ignore
    except Exception as e:  # pragma: no cover
        raise RuntimeError("Missing dependency: openpyxl. Please `pip install openpyxl`.") from e
    return Workbook, load_workbook, Font


def _wb_bytes(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_style(ws) -> None:
    _, _, Font = _openpyxl()
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _cell_dt(v: Any) -> datetime | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        return None


def build_manual_invoice_template_xlsx() -> bytes:
    Workbook, _, _ = _openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Free Invoice Import"
    ws.append(
        [
            "invoice_number",
            "issued_at",
            "due_at",
            "client_name",
            "phone",
            "address",
            "city",
            "zip_code",
            "currency",
            "tax_rate",
            "order_discount_amount",
            "shipping_amount",
            "line_type",
            "sku",
            "product_id",
            "description",
            "uom",
            "quantity",
            "unit_price",
            "discount_amount",
        ]
    )
    ws.append(
        [
            "ULFREE001",
            "",
            "",
            "Cho QTE",
            "",
            "",
            "",
            "",
            "USD",
            0,
            0,
            0,
            "FREE",
            "",
            "",
            "Dịch vụ đi hàng VN -> US",
            "Service",
            1,
            150,
            0,
        ]
    )
    ws.append(
        [
            "ULFREE001",
            "",
            "",
            "Cho QTE",
            "",
            "",
            "",
            "",
            "USD",
            0,
            0,
            0,
            "PRODUCT",
            "UL1628T001",
            "",
            "16x28 - White Hand Towel Premium",
            "Dozen",
            5,
            16,
            0,
        ]
    )
    _header_style(ws)
    widths = {
        "A": 18,
        "B": 20,
        "C": 20,
        "D": 24,
        "E": 16,
        "F": 24,
        "G": 16,
        "H": 12,
        "I": 10,
        "J": 10,
        "K": 18,
        "L": 16,
        "M": 12,
        "N": 16,
        "O": 12,
        "P": 30,
        "Q": 12,
        "R": 10,
        "S": 12,
        "T": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    return _wb_bytes(wb)


class ExcelImportError(ValueError):
    def __init__(self, errors: list[str]):
        super().__init__("\n".join(errors))
        self.errors = errors


def _read_rows_from_xlsx(file_bytes: bytes) -> tuple[list[str], list[list[Any]]]:
    _, load_workbook, _ = _openpyxl()
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ExcelImportError(["Empty workbook"])
    header = [(_cell_str(x) or "").strip().lower() for x in (rows[0] or [])]
    data_rows: list[list[Any]] = [list(r or []) for r in rows[1:]]
    return header, data_rows


def parse_manual_invoice_import_xlsx(
    file_bytes: bytes,
    *,
    products_by_sku: dict[str, Product],
    products_by_id: dict[int, Product],
) -> dict[str, Any]:
    header, rows = _read_rows_from_xlsx(file_bytes)
    errors: list[str] = []
    idx = {name: header.index(name) for name in header}

    if "description" not in idx:
        errors.append("Missing column: description")
    if "quantity" not in idx:
        errors.append("Missing column: quantity")
    if "uom" not in idx:
        errors.append("Missing column: uom")
    if "unit_price" not in idx:
        errors.append("Missing column: unit_price")
    if errors:
        raise ExcelImportError(errors)

    invoice_number: str | None = None
    issued_at: datetime | None = None
    due_at: datetime | None = None
    client_name: str | None = None
    phone: str | None = None
    address: str | None = None
    city: str | None = None
    zip_code: str | None = None
    currency: str | None = None
    tax_rate: float | None = None
    order_discount_amount: float | None = None
    shipping_amount: float | None = None

    lines: list[dict[str, Any]] = []
    for row_i, row in enumerate(rows, start=2):
        description = _cell_str(row[idx["description"]] if idx["description"] < len(row) else "")
        sku = _cell_str(row[idx["sku"]] if "sku" in idx and idx["sku"] < len(row) else "")
        product_id_raw = _cell_str(row[idx["product_id"]] if "product_id" in idx and idx["product_id"] < len(row) else "")
        quantity_raw = row[idx["quantity"]] if idx["quantity"] < len(row) else None
        unit_price_raw = row[idx["unit_price"]] if idx["unit_price"] < len(row) else None
        uom = _cell_str(row[idx["uom"]] if idx["uom"] < len(row) else "")
        discount_raw = row[idx["discount_amount"]] if "discount_amount" in idx and idx["discount_amount"] < len(row) else 0

        if not any([description, sku, product_id_raw, _cell_str(quantity_raw), _cell_str(unit_price_raw), uom]):
            continue

        def _capture_str(column: str, current: str | None) -> str | None:
            if column not in idx or idx[column] >= len(row):
                return current
            value = _cell_str(row[idx[column]])
            if not value:
                return current
            if current is None:
                return value
            if value != current:
                errors.append(f"Row {row_i}: {column} mismatch")
            return current

        invoice_number = _capture_str("invoice_number", invoice_number)
        client_name = _capture_str("client_name", client_name)
        phone = _capture_str("phone", phone)
        address = _capture_str("address", address)
        city = _capture_str("city", city)
        zip_code = _capture_str("zip_code", zip_code)
        currency = _capture_str("currency", currency)

        if "issued_at" in idx and idx["issued_at"] < len(row):
            dt = _cell_dt(row[idx["issued_at"]])
            if dt:
                if issued_at is None:
                    issued_at = dt
                elif dt != issued_at:
                    errors.append(f"Row {row_i}: issued_at mismatch")
        if "due_at" in idx and idx["due_at"] < len(row):
            dt = _cell_dt(row[idx["due_at"]])
            if dt:
                if due_at is None:
                    due_at = dt
                elif dt != due_at:
                    errors.append(f"Row {row_i}: due_at mismatch")

        numeric_headers: dict[str, float | None] = {
            "tax_rate": tax_rate,
            "order_discount_amount": order_discount_amount,
            "shipping_amount": shipping_amount,
        }
        for column, current in numeric_headers.items():
            if column not in idx or idx[column] >= len(row):
                continue
            raw = row[idx[column]]
            if raw in (None, ""):
                continue
            try:
                value = float(raw)
            except Exception:
                errors.append(f"Row {row_i}: invalid {column}")
                continue
            if current is None:
                numeric_headers[column] = value
            elif value != current:
                errors.append(f"Row {row_i}: {column} mismatch")
        tax_rate = numeric_headers["tax_rate"]
        order_discount_amount = numeric_headers["order_discount_amount"]
        shipping_amount = numeric_headers["shipping_amount"]

        try:
            quantity = int(quantity_raw)
        except Exception:
            errors.append(f"Row {row_i}: invalid quantity")
            continue
        if quantity <= 0:
            errors.append(f"Row {row_i}: quantity must be > 0")
            continue

        try:
            unit_price = float(unit_price_raw)
        except Exception:
            errors.append(f"Row {row_i}: invalid unit_price")
            continue
        if unit_price < 0:
            errors.append(f"Row {row_i}: unit_price must be >= 0")
            continue

        try:
            discount_amount = float(discount_raw or 0)
        except Exception:
            errors.append(f"Row {row_i}: invalid discount_amount")
            continue
        if discount_amount < 0:
            errors.append(f"Row {row_i}: discount_amount must be >= 0")
            continue

        line_type_text = _cell_str(row[idx["line_type"]] if "line_type" in idx and idx["line_type"] < len(row) else "")
        product: Product | None = None
        parsed_product_id: int | None = None

        if sku:
            product = products_by_sku.get(sku.upper())
            if product:
                parsed_product_id = product.id
        if product_id_raw:
            try:
                pid = int(float(product_id_raw))
            except Exception:
                errors.append(f"Row {row_i}: invalid product_id")
                continue
            product_from_id = products_by_id.get(pid)
            if product_from_id is None:
                errors.append(f"Row {row_i}: unknown product_id {pid}")
                continue
            if product is not None and product.id != product_from_id.id:
                errors.append(f"Row {row_i}: sku/product_id mismatch")
                continue
            product = product_from_id
            parsed_product_id = product.id

        if line_type_text:
            try:
                line_type = InvoiceLineType(line_type_text.upper())
            except Exception:
                errors.append(f"Row {row_i}: line_type must be PRODUCT or FREE")
                continue
        else:
            line_type = InvoiceLineType.PRODUCT if product is not None else InvoiceLineType.FREE

        if line_type == InvoiceLineType.PRODUCT and product is None:
            errors.append(f"Row {row_i}: PRODUCT line requires valid sku or product_id")
            continue

        line_description = description or (product.name if product else "")
        if not line_description:
            errors.append(f"Row {row_i}: description is required")
            continue
        if not uom:
            errors.append(f"Row {row_i}: uom is required")
            continue

        lines.append(
            {
                "id": None,
                "line_type": line_type.value,
                "product_id": parsed_product_id if line_type == InvoiceLineType.PRODUCT else None,
                "sku": sku or (product.sku if product else ""),
                "product_name": line_description,
                "uom": uom,
                "quantity": quantity,
                "unit_price": unit_price,
                "discount_amount": discount_amount,
            }
        )

    if not lines:
        errors.append("No invoice lines found")
    if not (client_name or "").strip():
        errors.append("client_name is required")
    if errors:
        raise ExcelImportError(errors)

    return {
        "invoice_number": invoice_number,
        "issued_at": issued_at.isoformat() if issued_at else None,
        "due_at": due_at.isoformat() if due_at else None,
        "client_name_snapshot": client_name or "",
        "tele_snapshot": phone or "",
        "address_snapshot": address or "",
        "city_snapshot": city or "",
        "zip_code_snapshot": zip_code or "",
        "currency": (currency or "USD").upper(),
        "tax_rate": tax_rate if tax_rate is not None else 0,
        "order_discount_amount": order_discount_amount if order_discount_amount is not None else 0,
        "shipping_amount": shipping_amount if shipping_amount is not None else 0,
        "lines": lines,
    }
