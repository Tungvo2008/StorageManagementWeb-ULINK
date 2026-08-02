from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path
import unittest

from openpyxl import Workbook, load_workbook

from app.services.amazon_shipment import (
    SolverBoxType,
    SolverSku,
    optimize_identical_cartons,
    parse_amazon_pack_csv,
    render_amazon_manifest_xlsx,
    render_amazon_pack_xlsx,
    render_amazon_pack_csv,
)


SAMPLE_CSV = '''\ufeff"Pack group number","1"
"Workflow name","workflow-test"
"SKUs","2"
"Units","29"

"SKU","Title","ASIN","FNSKU","UPC/EAN/ISBN/JAN/CODABAR","Condition","Prep type","Quantity","Box 1 units"
"AMZ-A","Product A","A000000001","X000000001","-","New","No prep needed","14","14"
"AMZ-B","Product B","A000000002","X000000002","-","New","No prep needed","15","15"


"","","","","","","","","","Box ID","To be assigned"
"","","","","","","","","","Box name","P1 - B1"
"","","","","","","","","","Box weight (lb):","20"
"","","","","","","","","","Box length (inch):","20"
"","","","","","","","","","Box width (inch):","16"
"","","","","","","","","","Box height (inch):","14"
'''


class AmazonShipmentCsvTests(unittest.TestCase):
    def test_parse_amazon_csv(self) -> None:
        parsed = parse_amazon_pack_csv(SAMPLE_CSV)
        self.assertEqual(parsed.pack_group_number, "1")
        self.assertEqual(parsed.workflow_name, "workflow-test")
        self.assertEqual(parsed.declared_unit_count, 29)
        self.assertEqual(len(parsed.item_rows), 2)
        self.assertEqual(len(parsed.boxes), 1)

    def test_export_repeats_identical_content_and_updates_totals(self) -> None:
        output = render_amazon_pack_csv(
            SAMPLE_CSV,
            per_box_quantities={"AMZ-A": 3, "AMZ-B": 3},
            boxes=[
                {
                    "name": f"P1 - B{index + 1}",
                    "weight_lb": Decimal("22.5"),
                    "length_in": Decimal("20"),
                    "width_in": Decimal("16"),
                    "height_in": Decimal("14"),
                }
                for index in range(5)
            ],
        )
        parsed = parse_amazon_pack_csv(output)
        self.assertEqual(parsed.declared_unit_count, 30)
        self.assertEqual(len(parsed.box_columns), 5)
        self.assertEqual([item.requested_quantity for item in parsed.item_rows], [15, 15])
        for item in parsed.item_rows:
            row = parsed.rows[item.row_index]
            self.assertEqual([row[column] for _, column in parsed.box_columns], ["3"] * 5)

    def test_manifest_export_fills_packaged_amazon_template(self) -> None:
        template_path = (
            Path(__file__).resolve().parents[1]
            / "assets"
            / "amazon"
            / "ManifestFileUpload_Template_MPL.xlsx"
        )
        output = render_amazon_manifest_xlsx(
            template_path.read_bytes(),
            items=[("AMZ-SKU-001", 12), ("AMZ-SKU-002", 7)],
        )
        workbook = load_workbook(BytesIO(output), read_only=True, data_only=False)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Instructions",
                "Data definitions",
                "Create workflow – template",
                "Create workflow – example",
            ],
        )
        sheet = workbook["Create workflow – template"]
        self.assertEqual(sheet["A9"].value, "AMZ-SKU-001")
        self.assertEqual(sheet["B9"].value, 12)
        self.assertEqual(sheet["A10"].value, "AMZ-SKU-002")
        self.assertEqual(sheet["B10"].value, 7)

    def test_pack_xlsx_export_fills_box_quantities_and_dimensions(self) -> None:
        source = Workbook()
        instructions = source.active
        instructions.title = "Instructions"
        sheet = source.create_sheet("Box packing information")
        source.create_sheet("Metadata")
        sheet["I3"] = "Total box count:"
        sheet["M3"] = 5
        headers = [
            "SKU", "Product title ", "Id", "ASIN", "FNSKU", "Condition",
            "Prep type", "Who preps units?", "Who labels units?",
            "Expected quantity", "Boxed quantity",
        ]
        for column, value in enumerate(headers, start=1):
            sheet.cell(5, column).value = value
        for offset in range(7):
            column = 13 + offset
            sheet.cell(5, column).value = f'=IF(M3>={offset + 1},"Box {offset + 1} quantity","")'
        sheet["A6"] = "AMZ-A"
        sheet["J6"] = 10
        sheet["K6"] = "=SUM(M6:Q6)"
        sheet["A7"] = "AMZ-B"
        sheet["J7"] = 15
        sheet["K7"] = "=SUM(M7:Q7)"
        sheet["A9"] = "Name of box"
        sheet["A10"] = "Box weight (lb):"
        sheet["A11"] = "Box width (inch):"
        sheet["A12"] = "Box length (inch):"
        sheet["A13"] = "Box height (inch):"
        buffer = BytesIO()
        source.save(buffer)

        output = render_amazon_pack_xlsx(
            buffer.getvalue(),
            per_box_quantities={"AMZ-A": 2, "AMZ-B": 3},
            boxes=[
                {
                    "name": f"P1 - B{index + 1}",
                    "weight_lb": 20 + index,
                    "length_in": 20,
                    "width_in": 16,
                    "height_in": 14,
                }
                for index in range(5)
            ],
        )
        workbook = load_workbook(BytesIO(output), data_only=False)
        filled = workbook["Box packing information"]
        self.assertEqual(filled["M3"].value, 5)
        self.assertEqual([filled.cell(6, column).value for column in range(13, 18)], [2] * 5)
        self.assertEqual([filled.cell(7, column).value for column in range(13, 18)], [3] * 5)
        self.assertEqual(filled["M9"].value, "P1 - B1")
        self.assertEqual(filled["M10"].value, 20)
        self.assertEqual(filled["M11"].value, 16)
        self.assertEqual(filled["M12"].value, 20)
        self.assertEqual(filled["M13"].value, 14)


class AmazonShipmentOptimizerTests(unittest.TestCase):
    def test_optimizer_rounds_to_identical_five_box_plan(self) -> None:
        skus = [
            SolverSku(
                amazon_sku="AMZ-A",
                title="Product A",
                requested_quantity=14,
                available_quantity=20,
                unit_weight_lb=1.0,
                capacities={1: 10},
            ),
            SolverSku(
                amazon_sku="AMZ-B",
                title="Product B",
                requested_quantity=15,
                available_quantity=20,
                unit_weight_lb=0.5,
                capacities={1: 10},
            ),
        ]
        plans = optimize_identical_cartons(
            skus=skus,
            box_types=[
                SolverBoxType(
                    id=1,
                    name="20x16x14",
                    length_in=20,
                    width_in=16,
                    height_in=14,
                    empty_weight_lb=1,
                    max_weight_lb=50,
                )
            ],
            min_box_count=5,
            max_box_count=8,
        )
        self.assertTrue(plans)
        recommended = plans[0]
        self.assertEqual(recommended["box_count"], 5)
        self.assertEqual(
            {item["amazon_sku"]: item["per_box_quantity"] for item in recommended["items"]},
            {"AMZ-A": 3, "AMZ-B": 3},
        )
        self.assertEqual(recommended["adjusted_unit_count"], 30)
        self.assertEqual(recommended["absolute_quantity_change"], 1)

    def test_optimizer_respects_fractional_capacity(self) -> None:
        skus = [
            SolverSku("A", "A", 15, 20, None, {1: 15}),
            SolverSku("B", "B", 30, 40, None, {1: 30}),
        ]
        plans = optimize_identical_cartons(
            skus=skus,
            box_types=[SolverBoxType(1, "Mixed", 20, 20, 20, 1, None)],
            min_box_count=5,
            max_box_count=5,
        )
        self.assertTrue(plans)
        recommended = plans[0]
        self.assertLessEqual(recommended["capacity_utilization"], 1.0)
        self.assertTrue(all(item["per_box_quantity"] >= 1 for item in recommended["items"]))


if __name__ == "__main__":
    unittest.main()
