from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any


HEADERS = [
    "sku",
    "name",
    "brand",
    "category",
    "unit_size",
    "case_pack",
    "wholesale_price",
    "currency",
    "country_of_origin",
    "upc",
    "stock_qty",
    "badges",
    "catalog_enabled",
    "is_active",
    "sort_order",
]


def build_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(HEADERS)
    sheet.append(["", "Sea Salt Potato Chips", "Coastal Harvest", "Snacks", "5 oz (142 g)", 12, 18.5, "USD", "USA", "", 100, "New", True, True, 10])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="15509B")
    widths = [20, 34, 22, 22, 18, 12, 18, 10, 20, 18, 12, 20, 18, 12, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_workbook(payload: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(payload), data_only=True, read_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        return []
    headers = [str(value or "").strip().lower() for value in raw_headers]
    missing = [header for header in ("name",) if header not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    parsed: list[dict[str, Any]] = []
    for row_number, values in enumerate(rows, start=2):
        item = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
        if not any(value not in (None, "") for value in item.values()):
            continue
        item["_row"] = row_number
        parsed.append(item)
    return parsed


def text(value: Any) -> str:
    return str(value or "").strip()


def integer(value: Any, default: int = 0) -> int:
    if value in (None, ""):
        return default
    return int(float(str(value).replace(",", "")))


def decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", ""))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid price: {value}") from exc


def boolean(value: Any, default: bool = True) -> bool:
    if value in (None, ""):
        return default
    return text(value).lower() in {"1", "true", "yes", "y", "active"}
